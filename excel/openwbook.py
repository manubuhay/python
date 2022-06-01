from openpyxl import load_workbook

your_workbook2 = load_workbook('Test.xlsx')
print(your_workbook2.sheetnames)

print(your_workbook2['Products']['A2'].value)

# But, what if I want to read the whole workbook?
# First, we would need to find out how many rows and columns the workbook has. 
# This can be achieved using .max_row and .max_column functions.
row = your_workbook2['Products'].max_row
col = your_workbook2['Products'].max_column

print(row)
print(col)

# Next, I would run two for loops; One would return every data on our rows, 
# and the other would return data in our columns.

for i in range(1, row+1):
    for j in range (1, col +1):
        print(your_workbook2['Products'].cell(i, j).value)

# Now, let’s move to cell formatting.
# Cell formatting customizes the look of the cell in the workbook. 
# We would perform simple formatting on our loaded test workbook. 
# But, first, we would create a worksheet object.

Wb = your_workbook2['Products']

# Then, we would import PatternFill.
# Pattern Fill allows us to change the colour and pattern of the cells. 
# To define our PatternFill, we need to pass a pattern type and fgColor, which is a hex code.

from openpyxl.styles import PatternFill

first_style = PatternFill(patternType = 'solid', fgColor = '00FF00') #00FF00 = Green
Wb['B7'].fill = first_style

# Since we are changing the workbook, we have to save it once done running our code. 
# So now, take a look at the spreadsheet. You will see the changes have been made to cell B7.
# To save the spreadsheet, use .save() function:

your_workbook2.save(filename='Test.xlsx') 