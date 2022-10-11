from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule

your_workbook2 = load_workbook('Book1.xlsx')
print(your_workbook2.sheetnames)
print(your_workbook2['Sheet1']['A2'].value)

row = your_workbook2['Sheet1'].max_row
col = your_workbook2['Sheet1'].max_column

print(row)
print(col)

# for i in range(1, row+1):
#     for j in range (1, col +1):
#         print(your_workbook2['Sheet1'].cell(i, j).value)

#Create worksheet object
bookObj = your_workbook2['Sheet1']
#Import patternfill
from openpyxl.styles import PatternFill

first_style = PatternFill(patternType = 'solid', fgColor = '00FF00')
bookObj['B7'].fill = first_style

your_workbook2.save(filename='Book1.xlsx')
# bookObj=your_workbook2.active
#
# Condition_style = ColorScaleRule(start_type = 'min', start_color = 'E0FFFF', end_type = 'max', end_color = '008080')
#
# row = bookObj.max_row
# col = bookObj.max_column
#
# for i in range(1,row+1):
#     if your_workbook2[i].value == "5":
#         Wb.conditional_formatting.add(your_workbook2[i], Condition_style)
#
# your_workbook2.save(filename='Test.xlsx')
